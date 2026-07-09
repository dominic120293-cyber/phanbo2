import pulp
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def run_allocation(input_file_path, output_file_path):
# ================== ĐỌC DỮ LIỆU TỪ FILE 1112.xlsx ==================
wb_data = openpyxl.load_workbook("1112.xlsx")
ws_data = wb_data.active

# Đọc danh sách STS từ dòng 1, cột J đến N (index 9-13)
sts_from_header = []
for col in range(9, 14):
    val = ws_data.cell(1, col).value
    if val:
        sts_from_header.append(val)

default_sts = ['STS01', 'STS02', 'STS03', 'STS04', 'STS05']
stss = []
for i in range(5):
    if i < len(sts_from_header):
        stss.append(sts_from_header[i])
    else:
        stss.append(default_sts[i])

print("Danh sách STS:", stss)

# --- Đọc supply (cột A: Yard, B-F: WC1-5) ---
supply = {}
for row in ws_data.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    yard = row[0]
    sup_dict = {}
    for wc_idx, val in enumerate(row[1:6], start=1):
        sup_dict[wc_idx] = int(val) if val is not None else 0
    supply[yard] = sup_dict

# --- Đọc demand (cột I: WC, cột J-N: STS) ---
demand = {}
for row in ws_data.iter_rows(min_row=2, values_only=True):
    wc_val = row[8]
    if wc_val is None:
        continue
    if isinstance(wc_val, str) and wc_val.startswith("WC"):
        wc = int(wc_val[2:])
    else:
        wc = int(wc_val)
    sts_dict = {}
    for i, sts in enumerate(stss):
        qty = row[9 + i]
        sts_dict[sts] = int(qty) if qty is not None else 0
    demand[wc] = sts_dict

yards = list(supply.keys())
wcs = sorted(demand.keys())

# ================== GIẢI MÔ HÌNH ==================
prob = pulp.LpProblem("Yard_Allocation_to_STS", pulp.LpMinimize)

x = pulp.LpVariable.dicts("assign", (yards, stss), cat='Binary')
alloc = pulp.LpVariable.dicts("alloc",
                             ((y, wc, s) for y in yards for wc in wcs for s in stss),
                             lowBound=0, cat='Continuous')

prob += pulp.lpSum(x[y][s] for y in yards for s in stss)

for y in yards:
    for wc in wcs:
        sup = supply[y].get(wc, 0)
        prob += pulp.lpSum(alloc[y, wc, s] for s in stss) == sup
        for s in stss:
            prob += alloc[y, wc, s] <= sup * x[y][s]

for wc in wcs:
    for s in stss:
        prob += pulp.lpSum(alloc[y, wc, s] for y in yards) == demand[wc].get(s, 0)

prob.solve(pulp.PULP_CBC_CMD(msg=False))

print("Status:", pulp.LpStatus[prob.status])
print("Tổng số cặp Yard-STS:", int(pulp.value(prob.objective)))

# ================== XUẤT FILE EXCEL VỚI 3 SHEET ==================
wb_out = openpyxl.Workbook()

# ---- 1. SHEET "Matrix" (tổng hợp) ----
ws_matrix = wb_out.active
ws_matrix.title = "Matrix"

data = defaultdict(lambda: defaultdict(float))
for y in yards:
    for s in stss:
        total = 0
        for wc in wcs:
            total += pulp.value(alloc[y, wc, s]) or 0
        if total > 0:
            data[s][y] = total

blocks_with_assign = sorted({y for s in data for y in data[s]})

def style_header(ws, row):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

row = 1
ws_matrix.cell(row, 1, "STS")
col = 2
for block in blocks_with_assign:
    ws_matrix.cell(row, col, block)
    col += 1
ws_matrix.cell(row, col, "Tổng STS")

style_header(ws_matrix, row)

sts_sorted = sorted(stss)
row = 2
for s in sts_sorted:
    ws_matrix.cell(row, 1, s)
    total_row = 0
    col = 2
    for block in blocks_with_assign:
        qty = data[s].get(block, 0)
        ws_matrix.cell(row, col, qty)
        total_row += qty
        col += 1
    ws_matrix.cell(row, col, total_row)
    row += 1

# Hàng tổng Block
ws_matrix.cell(row, 1, "Tổng Block")
col = 2
grand_total = 0
for block in blocks_with_assign:
    col_sum = sum(data[s].get(block, 0) for s in sts_sorted)
    ws_matrix.cell(row, col, col_sum)
    grand_total += col_sum
    col += 1
ws_matrix.cell(row, col, grand_total)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
for r in range(2, row + 1):
    for c in range(1, ws_matrix.max_column + 1):
        cell = ws_matrix.cell(r, c)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
for c in range(1, ws_matrix.max_column + 1):
    cell = ws_matrix.cell(row, c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

note_row = row + 2
ws_matrix.cell(note_row, 1, "Tổng số cặp Yard-STS đã dùng:")
ws_matrix.cell(note_row, 2, int(pulp.value(prob.objective)))
ws_matrix.cell(note_row, 1).font = Font(bold=True)

# ---- 2. SHEET "Detail" (chi tiết từng cặp Yard-STS theo WC) ----
ws_detail = wb_out.create_sheet("Detail")
detail_headers = ["Yard", "STS", "WC1", "WC2", "WC3", "WC4", "WC5", "Total"]
for col, header in enumerate(detail_headers, 1):
    cell = ws_detail.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

row = 2
for y in yards:
    for s in stss:
        wc_values = []
        total = 0
        for wc in wcs:
            qty = pulp.value(alloc[y, wc, s]) or 0
            wc_values.append(qty)
            total += qty
        if total > 0:
            ws_detail.cell(row, 1, y)
            ws_detail.cell(row, 2, s)
            for idx, qty in enumerate(wc_values, start=3):
                ws_detail.cell(row, idx, qty)
            ws_detail.cell(row, 8, total)
            for col in range(1, 9):
                cell = ws_detail.cell(row, col)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            row += 1

# ---- 3. SHEET "Combined" (ma trận chi tiết WC) ----
ws_comb = wb_out.create_sheet("Combined")

blocks = blocks_with_assign

# Header dòng 1 và 2
col = 2
for block in blocks:
    start_col = col
    end_col = col + 5
    ws_comb.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws_comb.cell(1, start_col, block)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    for wc_idx, wc_label in enumerate(["WC1", "WC2", "WC3", "WC4", "WC5", "Total"], start=start_col):
        cell = ws_comb.cell(2, wc_idx, wc_label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    col = end_col + 1

# Cột A: "STS" gộp 2 dòng
ws_comb.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
cell = ws_comb.cell(1, 1, "STS")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
cell.alignment = Alignment(horizontal="center")
cell.border = thin_border

total_columns = 1 + len(blocks) * 6
# Thêm cột tổng STS
ws_comb.merge_cells(start_row=1, start_column=total_columns + 1, end_row=2, end_column=total_columns + 1)
cell = ws_comb.cell(1, total_columns + 1, "Tổng STS")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
cell.alignment = Alignment(horizontal="center")
cell.border = thin_border

# Đổ dữ liệu
row = 3
for s in sts_sorted:
    ws_comb.cell(row, 1, s)
    col = 2
    total_row_all = 0
    for block in blocks:
        wc_vals = []
        total_block = 0
        for wc in wcs:
            qty = pulp.value(alloc[block, wc, s]) or 0
            wc_vals.append(qty)
            total_block += qty
        for idx, val in enumerate(wc_vals, start=col):
            ws_comb.cell(row, idx, val)
        ws_comb.cell(row, col + 5, total_block)
        total_row_all += total_block
        col += 6
    ws_comb.cell(row, total_columns + 1, total_row_all)
    row += 1

# Hàng tổng Block
row_sum = row
ws_comb.cell(row_sum, 1, "Tổng Block")
for col in range(2, total_columns + 2):
    sum_val = 0
    for r in range(3, row):
        val = ws_comb.cell(r, col).value
        if val is not None:
            sum_val += val
    ws_comb.cell(row_sum, col, sum_val)

# Định dạng tất cả ô (dữ liệu và header)
for r in range(1, row_sum + 1):
    for c in range(1, total_columns + 2):
        cell = ws_comb.cell(r, c)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if r >= 3:
            cell.border = thin_border
        if r == row_sum:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# ---- Tự động điều chỉnh độ rộng cột cho cả 3 sheet ----
for ws in [ws_matrix, ws_detail, ws_comb]:
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 20)

wb_out.save("Allocation_Result.xlsx")
print("✅ Đã xuất file: Allocation_Result.xlsx với 3 sheet: Matrix, Detail, Combined")
    pass
